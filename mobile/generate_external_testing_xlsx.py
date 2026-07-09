"""Generate BUDDY_WALK_EXTERNAL_TESTING.xlsx for TestFlight external testers.

Run:  python generate_external_testing_xlsx.py
Requires: openpyxl
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RUNS = 3
BUILD = "1.0.0 (3)"
FEEDBACK = "raymondsekyere99@gmail.com, dylansch7@gmail.com"
OUT = "exports/BUDDY_WALK_EXTERNAL_TESTING.xlsx"

ACCENT = "0D6E7A"
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color=ACCENT)
SECTION_FONT = Font(bold=True, size=12, color=ACCENT)
LABEL_FONT = Font(bold=True, size=10)
MUTED_FONT = Font(italic=True, size=10, color="555555")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER


def style_table(ws, start_row: int, end_row: int, cols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP


def set_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_readme_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Read Me", 0)
    ws.sheet_view.showGridLines = False
    set_widths(ws, {1: 92})

    lines = [
        ("Buddy Walk — External TestFlight Testing (Build 1.0.0 (3))", TITLE_FONT),
        ("", None),
        ("How to use", SECTION_FONT),
        ("1. Install Buddy Walk from the TestFlight link you received.", None),
        ("2. Fill in the Session sheet (tester info, pre-checklist, end summary).", None),
        ("3. Run each test category on the Test Runs sheet (3 runs each; Companion = 1 run).", None),
        ("4. Tap \"New Test\" on the main screen before each run to clear the screen and the AI's chat memory. (Fully closing the app or signing out also resets it.)", None),
        ("5. Log bugs on the Issues Log sheet as you find them.", None),
        ("", None),
        ("Feedback", SECTION_FONT),
        (f"Email: {FEEDBACK}", None),
        ("", None),
        ("What we are testing this round", SECTION_FONT),
        ("1A — Photo + Q&A (storefront / building sign)", None),
        ("1B — Hold-to-capture video on a fixed scene in front of you", None),
        ("1C — Voice: Tap to Ask (including a second question back-to-back)", None),
        ("2 — VoiceOver / accessibility on main flows (incl. video capture)", None),
        ("3 — Companion Mode: share a maps link once with a contact", None),
        ("4 — MTA subway arrival times (NYC only)", None),
        ("", None),
        ("System Metrics sheet", SECTION_FONT),
        ("Lists what Buddy Walk logs automatically vs what testers score manually.", MUTED_FONT),
        ("", None),
        ("Not in scope this round", SECTION_FONT),
        ("Turn-by-turn navigation, saved places, and route guidance.", MUTED_FONT),
        ("", None),
        ("Tips", SECTION_FONT),
        ("Use different storefronts, scenes, and subway lines each run when you can.", MUTED_FONT),
        ("Sign-in is skipped in this build — you should land on permissions, then the main screen.", MUTED_FONT),
        ("Companion sharing uses a Google Maps pin unless live tracking is available.", MUTED_FONT),
    ]
    for i, (text, font) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = WRAP


def build_session_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Session"
    ws.sheet_view.showGridLines = False
    set_widths(ws, {1: 24, 2: 30, 3: 14, 4: 14, 5: 14, 6: 14, 7: 36})

    ws["A1"] = "Buddy Walk — External TestFlight Session"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")

    ws["A2"] = f"Build: {BUILD}  |  Feedback: {FEEDBACK}"
    ws["A2"].font = MUTED_FONT
    ws.merge_cells("A2:G2")

    ws["A3"] = "Run each category on Test Runs (3 runs each; Companion = 1 run)."
    ws["A3"].font = MUTED_FONT
    ws.merge_cells("A3:G3")

    row = 5
    ws.cell(row=row, column=1, value="Session info").font = SECTION_FONT
    row += 1
    fields = [
        ("Tester name", ""),
        ("Date", ""),
        ("Device", "e.g. iPhone 15"),
        ("iOS version", "e.g. iOS 18"),
        ("Screen reader", "VoiceOver / Off"),
        ("City / area", "e.g. Manhattan"),
        ("TestFlight build", BUILD),
    ]
    for label, hint in fields:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = BORDER
        val = ws.cell(row=row, column=2, value=hint if hint.startswith("e.g.") or hint.startswith(BUILD[:3]) else "")
        val.border = BORDER
        val.alignment = WRAP
        if hint and not hint.startswith("e.g.") and hint != BUILD:
            val.font = MUTED_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Pre-session checklist").font = SECTION_FONT
    row += 1
    pre = [
        "P1 — Open Buddy Walk from TestFlight",
        "P2 — Finish permissions (location, camera, microphone)",
        "P3 — Land on main screen (camera, Ask field, Submit, Tap to Ask)",
        "P4 — Volume up or use headphones",
        "P5 — Confirm you are on build 1.0.0 (3) in TestFlight",
    ]
    ws.cell(row=row, column=1, value="Step").font = LABEL_FONT
    ws.cell(row=row, column=2, value="Done (Y/N)").font = LABEL_FONT
    style_header_row(ws, row, 2)
    row += 1
    for step in pre:
        ws.cell(row=row, column=1, value=step).border = BORDER
        ws.cell(row=row, column=2, value="").border = BORDER
        ws.cell(row=row, column=2).alignment = CENTER
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="End-of-session summary").font = SECTION_FONT
    row += 1
    headers = ["Test", "Runs logged", "Pass", "Fail", "Partial", "Overall"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    summary_tests = [
        ("1A Photo + Q&A", RUNS),
        ("1B Hold-to-capture video", RUNS),
        ("1C Voice (Tap to Ask)", RUNS),
        ("2 VoiceOver", RUNS),
        ("3 Companion share", 1),
        ("4 MTA arrival", RUNS),
    ]
    for test, run_count in summary_tests:
        row += 1
        ws.cell(row=row, column=1, value=test).border = BORDER
        ws.cell(row=row, column=2, value=f"/{run_count}").border = BORDER
        for c in range(3, 7):
            ws.cell(row=row, column=c, value="").border = BORDER

    row += 2
    ws.cell(row=row, column=1, value="Top 3 issues").font = LABEL_FONT
    for i in range(1, 4):
        row += 1
        ws.cell(row=row, column=1, value=f"{i}.").font = LABEL_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row=row, column=2, value="").border = BORDER

    row += 2
    for label in ("What worked best?", "What would block daily use?"):
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=7)
        ws.cell(row=row, column=1, value="").border = BORDER
        row += 3


def build_test_runs_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Test Runs")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {1: 10, 2: 30, 3: 6, 4: 34, 5: 12, 6: 12, 7: 12, 8: 12, 9: 40})

    ws["A1"] = "Test Runs — log 3 runs per category (Companion = 1 run)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")

    sections = [
        {
            "id": "1A",
            "name": "Photo — storefront Q&A",
            "question": 'Tap once on camera. Ask: "What business or building is this?" then Submit.',
            "subject_col": "Storefront / sign tried",
            "headers": ["Test", "Category", "Run", "Storefront / sign tried", "Pass", "Fail", "Partial", "", "Notes"],
            "subjects": [""] * RUNS,
        },
        {
            "id": "1B",
            "name": "Hold-to-capture video",
            "question": 'Hold camera 3–5 sec on a fixed scene. Ask: "Describe what you see in front of me."',
            "subject_col": "Scene / subject tried",
            "headers": ["Test", "Category", "Run", "Scene / subject tried", "Pass", "Fail", "Partial", "Skipped", "Notes"],
            "subjects": [
                "Building façade or entrance",
                "Street sign or bus stop",
                "Storefront with visible sign",
            ],
        },
        {
            "id": "1C",
            "name": "Voice — Tap to Ask",
            "question": "Tap to Ask → speak → Submit. Run 3 is a second question immediately after the first.",
            "subject_col": "Question spoken",
            "headers": ["Test", "Category", "Run", "Question spoken", "Pass", "Fail", "Partial", "2nd Q worked?", "Notes"],
            "subjects": [
                "What street am I on?",
                "What intersection am I at?",
                "Second question right after first (any short question)",
            ],
        },
        {
            "id": "2",
            "name": "VoiceOver / accessibility",
            "question": "VoiceOver on — complete each flow using spoken labels only.",
            "subject_col": "Flow repeated",
            "headers": ["Test", "Category", "Run", "Flow repeated", "Pass", "Fail", "Partial", "", "Notes"],
            "subjects": [
                "Main screen buttons (Submit, Tap to Ask)",
                "Test 1A photo capture",
                "Test 1B hold-to-capture video",
            ],
        },
        {
            "id": "3",
            "name": "Companion Mode share",
            "runs": 1,
            "question": "Start sharing → share link → contact opens Google Maps pin (re-share after moving).",
            "subject_col": "Contact / second device",
            "headers": ["Test", "Category", "Run", "Contact / second device", "Pass", "Fail", "Partial", "Maps pin correct?", "Notes"],
            "subjects": [""],
        },
        {
            "id": "4",
            "name": "MTA subway arrival",
            "question": 'Ask: "When is the next [LINE] train arriving?" (NYC only)',
            "subject_col": "Line asked",
            "headers": ["Test", "Category", "Run", "Line asked", "Minutes plausible?", "Pass", "Fail", "Partial", "Notes"],
            "subjects": ["1", "2", "3"],
        },
    ]

    row = 3
    for sec in sections:
        ws.cell(row=row, column=1, value=f"Test {sec['id']} — {sec['name']}").font = SECTION_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
        ws.cell(row=row, column=1, value=sec["question"]).font = MUTED_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1

        headers = sec["headers"][:9]
        for i, h in enumerate(headers, 1):
            if h:
                ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, 9)
        row += 1

        subjects = sec["subjects"]
        section_runs = sec.get("runs", RUNS)
        for run in range(1, section_runs + 1):
            ws.cell(row=row, column=1, value=sec["id"]).border = BORDER
            ws.cell(row=row, column=2, value=sec["name"]).border = BORDER
            ws.cell(row=row, column=3, value=run).border = BORDER
            ws.cell(row=row, column=3).alignment = CENTER

            subject = subjects[run - 1] if run - 1 < len(subjects) else ""
            ws.cell(row=row, column=4, value=subject).border = BORDER

            if sec["id"] == "4":
                result_cols = (6, 7, 8)
                ws.cell(row=row, column=5, value="").border = BORDER
                ws.cell(row=row, column=5).alignment = CENTER
                notes_col = 9
            elif sec["id"] == "1B":
                result_cols = (5, 6, 7)
                ws.cell(row=row, column=8, value="").border = BORDER
                ws.cell(row=row, column=8).alignment = CENTER
                notes_col = 9
            elif sec["id"] in ("1C", "3"):
                result_cols = (5, 6, 7)
                ws.cell(row=row, column=8, value="").border = BORDER
                ws.cell(row=row, column=8).alignment = CENTER
                notes_col = 9
            else:
                result_cols = (5, 6, 7)
                notes_col = 9

            for c in result_cols:
                cell = ws.cell(row=row, column=c, value="")
                cell.border = BORDER
                cell.alignment = CENTER

            notes = ws.cell(row=row, column=notes_col, value="")
            notes.border = BORDER
            notes.alignment = WRAP
            style_table(ws, row, row, 9)
            row += 1

        row += 1

    ws.freeze_panes = "A4"


def build_issues_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Issues Log")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {1: 6, 2: 10, 3: 12, 4: 14, 5: 40, 6: 36, 7: 16, 8: 14, 9: 12})

    ws["A1"] = "External Issues Log"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = f"Build {BUILD} — log bugs found during TestFlight testing."
    ws["A2"].font = MUTED_FONT
    ws.merge_cells("A2:I2")

    headers = [
        "#",
        "Test",
        "Severity",
        "Platform",
        "Description",
        "Steps to reproduce",
        "Device / iOS",
        "Status",
        "Owner",
    ]
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))

    sev_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    status_dv = DataValidation(
        type="list",
        formula1='"Open,In Progress,Fixed,Won\'t fix,Duplicate"',
        allow_blank=True,
    )
    ws.add_data_validation(sev_dv)
    ws.add_data_validation(status_dv)

    for i in range(1, 51):
        row += 1
        ws.cell(row=row, column=1, value=i).border = BORDER
        ws.cell(row=row, column=1).alignment = CENTER
        for c in range(2, len(headers) + 1):
            ws.cell(row=row, column=c, value="").border = BORDER
            ws.cell(row=row, column=c).alignment = WRAP
        sev_dv.add(f"C{row}")
        status_dv.add(f"H{row}")

    ws.freeze_panes = "A5"


def build_system_metrics_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("System Metrics")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {1: 32, 2: 14, 3: 18, 4: 52})

    ws["A1"] = "Buddy Walk System Metrics"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = "Logged automatically in TestFlight builds (unless analytics opted out). No manual entry needed."
    ws["A2"].font = MUTED_FONT
    ws.merge_cells("A2:D2")

    headers = ["Metric", "Auto-log?", "Telemetry / source", "Notes"]
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    row += 1

    metrics = [
        (
            "Instruction accuracy",
            "Partial",
            "answer_rated, feedback_submitted",
            "Thumbs on answers + optional feedback text; no ground-truth route audit yet.",
        ),
        (
            "Instruction timing",
            "Partial",
            "navigation_started → arrived/stopped",
            "Session duration inferred from nav events; per-step speak timing not exported yet.",
        ),
        (
            "Instruction latency",
            "Yes",
            "answer_received.props.latencyMs",
            "Question submit → spoken answer ready (ms).",
        ),
        (
            "GPS localization error",
            "Partial",
            "coords.accuracy on Q&A requests",
            "Horizontal accuracy (m) sent with each question when location available.",
        ),
        (
            "Scene / object Q&A success",
            "Yes",
            "answer_received vs answer_failed / answer_rejected",
            "Photo, video, and text Q&A outcomes by feature tag.",
        ),
        (
            "Missed scene description rate",
            "No",
            "—",
            "Requires human labels or tester rubric (see Test Runs sheet).",
        ),
        (
            "False / hallucinated description rate",
            "No",
            "—",
            "Requires human labels or tester rubric (see Test Runs sheet).",
        ),
        (
            "Frequency of rerouting",
            "Yes",
            "navigation_off_route",
            "Counted when user leaves the active route corridor.",
        ),
        (
            "Number of AI conversations",
            "Yes",
            "question_asked (count per install/session)",
            "Each submitted question; view via /api/telemetry/summary.",
        ),
        (
            "Voice command recognition accuracy",
            "Partial",
            "voice_started, voice_stopped, question_asked",
            "Transcript stored locally until submit; no auto WER vs spoken audio yet.",
        ),
        (
            "Photo capture success",
            "Yes",
            "photo_captured",
            "Fires when a still frame is captured.",
        ),
        (
            "Video / hold-to-capture success",
            "Yes",
            "video_recorded",
            "Fires when hold-to-capture completes.",
        ),
        (
            "Companion share events",
            "Yes",
            "companion_session_created, companion_link_shared",
            "Maps-pin or live share flow started and link shared.",
        ),
    ]

    for metric, auto, source, notes in metrics:
        ws.cell(row=row, column=1, value=metric).border = BORDER
        ws.cell(row=row, column=2, value=auto).border = BORDER
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=3, value=source).border = BORDER
        ws.cell(row=row, column=4, value=notes).border = BORDER
        for c in range(1, 5):
            ws.cell(row=row, column=c).alignment = WRAP
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Dashboard").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Team: GET /api/telemetry/summary and /api/telemetry/events.csv on buddywalk.app (admin).")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1).font = MUTED_FONT
    ws.cell(row=row, column=1).alignment = WRAP

    ws.freeze_panes = "A5"


def main() -> None:
    wb = Workbook()
    build_readme_sheet(wb)
    build_session_sheet(wb)
    build_test_runs_sheet(wb)
    build_system_metrics_sheet(wb)
    build_issues_sheet(wb)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
