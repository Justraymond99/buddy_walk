"""Generate BUDDY_WALK_INTERNAL_TESTING.xlsx for internal QA sessions.

Run:  python generate_internal_testing_xlsx.py
Requires: openpyxl
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RUNS = 5
WEB_URL = "https://buddy-walk-mobile.vercel.app"
FEEDBACK = "raymondsekyere99@gmail.com, dylansch7@gmail.com"
OUT = "exports/BUDDY_WALK_INTERNAL_TESTING.xlsx"

# Styles
ACCENT = "0D6E7A"
ACCENT_LIGHT = "EEF4F5"
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color=ACCENT)
SECTION_FONT = Font(bold=True, size=12, color=ACCENT)
LABEL_FONT = Font(bold=True, size=10)
MUTED_FONT = Font(italic=True, size=10, color="555555")
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = CENTER


def style_table(ws, start_row: int, end_row: int, cols: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = WRAP


def add_result_validation(ws, col: int, start_row: int, end_row: int) -> None:
    col_letter = get_column_letter(col)
    dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Partial,Skipped,N/A"',
        allow_blank=True,
    )
    dv.error = "Choose Pass, Fail, Partial, Skipped, or N/A"
    dv.prompt = "Select result"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def set_widths(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_session_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Session"
    ws.sheet_view.showGridLines = False

    set_widths(ws, {1: 22, 2: 28, 3: 14, 4: 14, 5: 14, 6: 14, 7: 36})

    ws["A1"] = "Buddy Walk — Internal Testing Session"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")

    ws["A2"] = f"Web beta: {WEB_URL}  |  Feedback: {FEEDBACK}"
    ws["A2"].font = MUTED_FONT
    ws.merge_cells("A2:G2")

    ws["A3"] = "Run each test category 5 times. Log every run in the Test Runs sheet."
    ws["A3"].font = MUTED_FONT
    ws.merge_cells("A3:G3")

    row = 5
    ws.cell(row=row, column=1, value="Session info").font = SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    fields = [
        ("Tester name", ""),
        ("Date", ""),
        ("Device", "e.g. iPhone 14, Safari"),
        ("Platform", "Web / TestFlight / native"),
        ("Screen reader", "VoiceOver / TalkBack / Off"),
        ("Location", "Street / intersection, city"),
        ("Session lead", ""),
        ("App / build version", ""),
        ("Backend", "buddywalk.app"),
    ]
    row += 1
    for label, hint in fields:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=1).border = BORDER
        val = ws.cell(row=row, column=2, value=hint if hint and hint.startswith("e.g.") else "")
        val.border = BORDER
        val.alignment = WRAP
        if hint and not hint.startswith("e.g."):
            val.font = MUTED_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Pre-session checklist").font = SECTION_FONT
    row += 1
    pre = [
        "P1 — Finish permissions (location, camera, microphone)",
        "P2 — Land on main screen (camera, Ask field, Submit)",
        "P3 — Headphones on or volume up",
        "P4 — Phone browser for web beta (Safari iOS / Chrome Android)",
        "P5 — Note if location looks approximate (desktop Wi‑Fi)",
    ]
    ws.cell(row=row, column=1, value="Step").font = LABEL_FONT
    ws.cell(row=row, column=2, value="Done (Y/N)").font = LABEL_FONT
    style_header_row(ws, row, 2)
    row += 1
    for step in pre:
        ws.cell(row=row, column=1, value=step).border = BORDER
        c = ws.cell(row=row, column=2, value="")
        c.border = BORDER
        c.alignment = CENTER
        row += 1

    add_result_validation(ws, 2, row - len(pre), row - 1)

    row += 2
    ws.cell(row=row, column=1, value="Pre-session notes").font = LABEL_FONT
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=7)
    note = ws.cell(row=row, column=1, value="")
    note.border = BORDER
    note.alignment = WRAP

    row += 4
    ws.cell(row=row, column=1, value="End-of-session summary").font = SECTION_FONT
    row += 1
    headers = ["Test", "Runs logged", "Pass", "Fail", "Partial", "Overall"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))
    summary_tests = [
        "1A Photo + Q&A",
        "1B Video + Q&A",
        "1C Voice input",
        "2 Navigation",
        "3 VoiceOver",
        "4 Companion",
        "5 Saved Places",
        "6 MTA arrival",
    ]
    for test in summary_tests:
        row += 1
        ws.cell(row=row, column=1, value=test).border = BORDER
        ws.cell(row=row, column=2, value=f"/{RUNS}").border = BORDER
        for c in range(3, 7):
            ws.cell(row=row, column=c, value="").border = BORDER

    row += 2
    ws.cell(row=row, column=1, value="Top 3 issues").font = LABEL_FONT
    for i in range(1, 4):
        row += 1
        ws.cell(row=row, column=1, value=f"{i}.").font = LABEL_FONT
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws.cell(row=row, column=2, value="").border = BORDER

    row += 2
    for label in ("What worked best?", "What would block daily use?"):
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=7)
        ws.cell(row=row, column=1, value="").border = BORDER
        row += 3

    ws.cell(row=row, column=1, value="Recommend to another blind/low-vision traveler?").font = LABEL_FONT
    ws.cell(row=row, column=2, value="Yes / Maybe / No — because:")
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
    ws.cell(row=row, column=3, value="").border = BORDER


def build_test_runs_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Test Runs")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {1: 10, 2: 28, 3: 6, 4: 32, 5: 12, 6: 12, 7: 12, 8: 12, 9: 40})

    ws["A1"] = "Test Runs — log 5 runs per category"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")

    sections = [
        {
            "id": "1A",
            "name": "Photo — storefront",
            "question": 'Ask: "What business or building is this?"',
            "subject_col": "Storefront / location tried",
            "extra_cols": [],
            "subjects": [""] * RUNS,
        },
        {
            "id": "1B",
            "name": "Video — landmark",
            "question": 'Hold 3–5 sec. Ask: "Describe what you see in front of me."',
            "subject_col": "Landmark tried",
            "extra_cols": ["Skipped"],
            "subjects": [""] * RUNS,
        },
        {
            "id": "1C",
            "name": "Voice — Tap to Ask",
            "question": "Tap to Ask, then Submit",
            "subject_col": "Question spoken",
            "extra_cols": [],
            "subjects": [
                "What street am I on?",
                "What intersection am I at?",
                "What's near me?",
                "What businesses are around me?",
                "Am I facing north or south?",
            ],
        },
        {
            "id": "2",
            "name": "Hands-off navigation",
            "question": 'Ask: "How do I get to [destination]?" — auto-start, walk ≥2 min',
            "subject_col": "Destination",
            "extra_cols": [],
            "subjects": [""] * RUNS,
        },
        {
            "id": "3",
            "name": "VoiceOver / accessibility",
            "question": "Screen reader on — repeat flows without sight",
            "subject_col": "Flow repeated",
            "extra_cols": [],
            "subjects": [
                "Test 1A (storefront photo)",
                "Test 2 (navigation, ≥2 min or 2 steps)",
                "Submit + Tap to Ask labels only",
                "Test 1B (landmark video)",
                "Companion Mode controls",
            ],
        },
        {
            "id": "4",
            "name": "Companion Mode",
            "question": "Create session, share link, walk 1 block — location updates?",
            "subject_col": "Contact / second device",
            "extra_cols": [],
            "subjects": [""] * RUNS,
        },
        {
            "id": "5",
            "name": "Saved Places",
            "question": "Save alias, ask for directions by name",
            "subject_col": "Alias saved",
            "extra_cols": ["Question asked"],
            "subjects": ["test-home", "work", "home", "", ""],
            "questions": [
                "How do I get to test-home?",
                "How do I get to work?",
                "How do I get home?",
                "",
                "",
            ],
        },
        {
            "id": "6",
            "name": "MTA subway arrival (NYC)",
            "question": 'Ask: "When is the next [LINE] train arriving?"',
            "subject_col": "Line asked",
            "extra_cols": ["Station in answer"],
            "subjects": [""] * RUNS,
        },
    ]

    row = 3
    result_start = row

    for sec in sections:
        ws.cell(row=row, column=1, value=f"Test {sec['id']} — {sec['name']}").font = SECTION_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1
        ws.cell(row=row, column=1, value=sec["question"]).font = MUTED_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1

        headers = ["Test", "Category", "Run", sec["subject_col"]]
        headers.extend(sec["extra_cols"])
        headers.extend(["Pass", "Fail", "Partial"])
        if "Skipped" in sec["extra_cols"]:
            pass  # Skipped is data column not result
        headers.append("Notes")

        # Normalize headers for 1B (Skipped is per-run flag, not header duplicate)
        if sec["id"] == "1B":
            headers = ["Test", "Category", "Run", sec["subject_col"], "Pass", "Fail", "Partial", "Skipped", "Notes"]
        elif sec["id"] == "5":
            headers = ["Test", "Category", "Run", "Alias saved", "Question asked", "Pass", "Fail", "Partial", "Notes"]
        elif sec["id"] == "6":
            headers = ["Test", "Category", "Run", "Line asked", "Station in answer", "Pass", "Fail", "Partial", "Notes"]
        else:
            headers = ["Test", "Category", "Run", sec["subject_col"], "Pass", "Fail", "Partial", "Notes", ""]

        # Pad to 9 columns
        while len(headers) < 9:
            headers.append("")
        headers = headers[:9]

        for i, h in enumerate(headers, 1):
            if h:
                ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, 9)
        header_row = row
        row += 1

        for run in range(1, RUNS + 1):
            ws.cell(row=row, column=1, value=sec["id"]).border = BORDER
            ws.cell(row=row, column=2, value=sec["name"]).border = BORDER
            ws.cell(row=row, column=3, value=run).border = BORDER
            ws.cell(row=row, column=3).alignment = CENTER

            subject = sec["subjects"][run - 1] if run - 1 < len(sec["subjects"]) else ""
            ws.cell(row=row, column=4, value=subject).border = BORDER

            col = 5
            if sec["id"] == "5":
                q = sec.get("questions", [""] * RUNS)[run - 1]
                ws.cell(row=row, column=5, value=q).border = BORDER
                result_cols = (6, 7, 8)
                notes_col = 9
            elif sec["id"] == "6":
                ws.cell(row=row, column=5, value="").border = BORDER
                result_cols = (6, 7, 8)
                notes_col = 9
            elif sec["id"] == "1B":
                result_cols = (5, 6, 7)
                notes_col = 9
                ws.cell(row=row, column=8, value="").border = BORDER
                ws.cell(row=row, column=8).alignment = CENTER
            else:
                result_cols = (5, 6, 7)
                notes_col = 8 if sec["id"] != "1B" else 9

            for c in result_cols:
                cell = ws.cell(row=row, column=c, value="")
                cell.border = BORDER
                cell.alignment = CENTER

            notes = ws.cell(row=row, column=notes_col, value="")
            notes.border = BORDER
            notes.alignment = WRAP

            style_table(ws, row, row, 9)
            row += 1

        row += 1  # blank row between sections

    ws.freeze_panes = "A4"


def build_issues_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Issues Log")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {1: 6, 2: 10, 3: 12, 4: 14, 5: 40, 6: 36, 7: 16, 8: 14, 9: 12})

    ws["A1"] = "Internal Issues Log"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = "Log bugs and UX issues found during internal testing."
    ws["A2"].font = MUTED_FONT
    ws.merge_cells("A2:I2")

    headers = [
        "#",
        "Test",
        "Severity",
        "Platform",
        "Description",
        "Steps to reproduce",
        "Device / browser",
        "Status",
        "Owner",
    ]
    row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header_row(ws, row, len(headers))

    sev_dv = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
    status_dv = DataValidation(
        type="list",
        formula1='"Open,In Progress,Fixed,Won\'t fix,Duplicate"',
        allow_blank=True,
    )
    ws.add_data_validation(sev_dv)
    ws.add_data_validation(status_dv)

    for i in range(1, 51):
        row += 1
        ws.cell(row=row, column=1, value=i).border = BORDER
        ws.cell(row=row, column=1).alignment = CENTER
        for c in range(2, len(headers) + 1):
            ws.cell(row=row, column=c, value="").border = BORDER
            ws.cell(row=row, column=c).alignment = WRAP
        sev_dv.add(f"C{row}")
        status_dv.add(f"H{row}")

    ws.freeze_panes = "A5"


def build_readme_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Read Me", 0)
    ws.sheet_view.showGridLines = False
    set_widths(ws, {1: 90})

    lines = [
        ("Buddy Walk — Internal Testing Workbook", TITLE_FONT),
        ("", None),
        ("How to use", SECTION_FONT),
        ("1. Fill in the Session sheet (tester info, pre-checklist, end summary).", None),
        ("2. Run each test category 5 times on the Test Runs sheet.", None),
        ("3. Log bugs on the Issues Log sheet as you find them.", None),
        ("", None),
        ("Links", SECTION_FONT),
        (f"Web beta: {WEB_URL}", None),
        (f"Feedback: {FEEDBACK}", None),
        ("", None),
        ("Test categories (5 runs each)", SECTION_FONT),
        ("1A — Photo storefront + Q&A", None),
        ("1B — Video / hold landmark", None),
        ("1C — Voice Tap to Ask", None),
        ("2 — Hands-off navigation (native = haptics + shake; web = spoken only)", None),
        ("3 — VoiceOver / TalkBack", None),
        ("4 — Companion Mode live share", None),
        ("5 — Saved Places aliases", None),
        ("6 — MTA subway arrival (NYC only)", None),
        ("", None),
        ("Pass criteria quick reference", SECTION_FONT),
        ("Balanced testing: use different storefronts, landmarks, destinations, and subway lines each run.", MUTED_FONT),
        ("Native TestFlight: full haptic nav + shake-to-stop.", MUTED_FONT),
        ("Web: no haptics; use Stop Navigation button.", MUTED_FONT),
    ]
    for i, (text, font) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=text)
        if font:
            cell.font = font
        cell.alignment = WRAP


def main() -> None:
    wb = Workbook()
    build_readme_sheet(wb)
    build_session_sheet(wb)
    build_test_runs_sheet(wb)
    build_issues_sheet(wb)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
