"""Fill Issues Log (and optional session fields) in William Lu's testing workbook."""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path

DEFAULT_XLSX = Path(r"c:\Users\CISUSER\Downloads\William Lu - 6_18_2026 (1).xlsx")
DEFAULT_ODS = Path(r"c:\Users\CISUSER\Desktop\William Lu - 6_18_2026.xlsx.ods")

ISSUE_ROWS: list[dict[str, str]] = [
    {
        "num": "1",
        "test": "1C",
        "severity": "High",
        "platform": "TestFlight",
        "description": "Tap to Ask / voice input did not record or transcribe",
        "steps": "Tap to Ask → speak question → text field stays empty",
        "device": "iPhone 17, TestFlight",
        "status": "Open",
        "owner": "Engineering",
    },
    {
        "num": "2",
        "test": "1B",
        "severity": "Medium",
        "platform": "TestFlight",
        "description": "Hold-to-capture video/frames feels buggy",
        "steps": "Hold camera 3–5s on landmark → intermittent capture failures",
        "device": "iPhone 17, TestFlight",
        "status": "Open",
        "owner": "Engineering",
    },
    {
        "num": "3",
        "test": "2",
        "severity": "High",
        "platform": "TestFlight",
        "description": "Auto-navigation spoken directions feel inaccurate",
        "steps": 'Ask "How do I get to [destination]?" → walk → turn cues do not match street',
        "device": "iPhone 17, TestFlight",
        "status": "Open",
        "owner": "Engineering",
    },
    {
        "num": "4",
        "test": "5",
        "severity": "High",
        "platform": "TestFlight",
        "description": "Saved place resolves to wrong address; delete hard to find",
        "steps": "Save alias → ask directions → route to wrong block; Remove button unclear in VoiceOver",
        "device": "iPhone 17, TestFlight",
        "status": "In progress",
        "owner": "Engineering",
    },
    {
        "num": "5",
        "test": "4",
        "severity": "Critical",
        "platform": "TestFlight",
        "description": "Companion Mode share link does not work reliably",
        "steps": "Create session → share link → recipient opens link → map/location fails",
        "device": "iPhone 17 + second devices",
        "status": "In progress",
        "owner": "Engineering",
    },
    {
        "num": "6",
        "test": "6",
        "severity": "Critical",
        "platform": "TestFlight",
        "description": "MTA arrival times appear ~4 hours in the future",
        "steps": 'Ask "When is the next R train arriving?" at 11:02 → answer ~3:04 PM',
        "device": "iPhone 17, TestFlight, Brooklyn",
        "status": "Fixed",
        "owner": "Engineering",
    },
    {
        "num": "7",
        "test": "All",
        "severity": "Medium",
        "platform": "TestFlight",
        "description": "Spoken AI answers hard to hear / volume too low",
        "steps": "Submit any question → TTS playback quiet or routed to earpiece",
        "device": "iPhone 17, headphones",
        "status": "Fixed",
        "owner": "Engineering",
    },
    {
        "num": "8",
        "test": "5",
        "severity": "Medium",
        "platform": "TestFlight",
        "description": "Saved-place routing prefers long walk over nearer subway",
        "steps": "Save gym/home → ask directions → suggests Atlantic Ave walk vs nearer station",
        "device": "iPhone 17, TestFlight",
        "status": "Open",
        "owner": "Engineering",
    },
]

SESSION_FIELDS: dict[str, str] = {
    "Platform": "TestFlight (native iOS)",
    "Screen reader": "VoiceOver (Test 3 only)",
    "Location": "Downtown Brooklyn / Greenwich St area, NYC",
    "Session lead": "Raymond",
    "App / build version": "TestFlight build 2+",
    "Backend": "buddywalk.app",
}


def issue_to_xlsx_columns(issue: dict[str, str]) -> dict[int, str]:
    return {
        2: issue["test"],
        3: issue["severity"],
        4: issue["platform"],
        5: issue["description"],
        6: issue["steps"],
        7: issue["device"],
        8: issue["status"],
        9: issue["owner"],
    }


def issue_to_ods_columns(issue: dict[str, str]) -> dict[int, str]:
    return {
        1: issue["test"],
        2: issue["severity"],
        3: issue["platform"],
        4: issue["description"],
        5: issue["steps"],
        6: issue["device"],
        7: issue["status"],
        8: issue["owner"],
    }


def patch_issues_xlsx(ws) -> None:
    for issue in ISSUE_ROWS:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell_val = row[0].value
            if cell_val is None:
                continue
            normalized = str(cell_val).strip().removesuffix(".0")
            if normalized == issue["num"]:
                for col_idx, value in issue_to_xlsx_columns(issue).items():
                    row[col_idx - 1].value = value
                break


def patch_session_xlsx(ws) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        label = row[0].value
        if label is None:
            continue
        key = str(label).strip()
        if key in SESSION_FIELDS:
            row[1].value = SESSION_FIELDS[key]


def patch_xlsx(path: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path)
    if "Issues Log" in wb.sheetnames:
        patch_issues_xlsx(wb["Issues Log"])
    if "Session" in wb.sheetnames:
        patch_session_xlsx(wb["Session"])
    wb.save(path)


def cell_text(cell) -> str:
    from odf.text import P

    return "".join(
        "".join(getattr(node, "data", "") for node in p.childNodes)
        for p in cell.getElementsByType(P)
    )


def set_cell_text(cell, text: str) -> None:
    from odf.text import P

    for p in cell.getElementsByType(P):
        cell.removeChild(p)
    p = P()
    p.addText(text)
    cell.addElement(p)


def expand_row(row) -> list[str]:
    from odf.table import TableCell

    values: list[str] = []
    for cell in row.getElementsByType(TableCell):
        reps = int(cell.getAttribute("numbercolumnsrepeated") or 1)
        text = cell_text(cell)
        for i in range(reps):
            values.append(text if i == 0 else "")
    return values


def rebuild_row(row, values: list[str]) -> None:
    from odf.table import TableCell, TableRow

    cells = list(row.getElementsByType(TableCell))
    padding = None
    if cells and int(cells[-1].getAttribute("numbercolumnsrepeated") or 1) > 100:
        padding = cells[-1]

    for cell in cells:
        if cell is not padding:
            row.removeChild(cell)

    for value in values:
        cell = TableCell()
        if value:
            set_cell_text(cell, value)
        row.addElement(cell)

    if padding is not None:
        row.addElement(padding)


def set_row_values(row, updates: dict[int, str]) -> None:
    values = expand_row(row)
    keep = max(max(updates) + 1 if updates else 1, len(values) if len(values) < 50 else 0)
    if len(values) >= 50:
        keep = max(keep, max(updates) + 1 if updates else 9)
    values = (values + [""] * keep)[:keep]
    for idx, value in updates.items():
        values[idx] = value
    rebuild_row(row, values)


def find_data_row(sheet, match: tuple[str, ...]):
    from odf.table import TableRow

    for row in sheet.getElementsByType(TableRow):
        vals = expand_row(row)
        if vals[: len(match)] == list(match):
            return row
    return None


def patch_ods(path: Path) -> None:
    from odf.opendocument import load
    from odf.table import Table

    doc = load(str(path))
    sheets = {t.getAttribute("name"): t for t in doc.spreadsheet.getElementsByType(Table)}

    for issue in ISSUE_ROWS:
        row = find_data_row(sheets["Issues Log"], (issue["num"],))
        if row:
            set_row_values(row, issue_to_ods_columns(issue))

    for key, value in SESSION_FIELDS.items():
        row = find_data_row(sheets["Session"], (key,))
        if row:
            set_row_values(row, {1: value})

    doc.save(str(path))


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        help="Path to .xlsx or .ods workbook (defaults to Downloads xlsx)",
    )
    args = parser.parse_args()

    path = args.workbook or DEFAULT_XLSX
    if not path.exists() and path == DEFAULT_XLSX and DEFAULT_ODS.exists():
        path = DEFAULT_ODS

    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    backup = path.with_suffix(path.suffix + ".bak")
    if not backup.exists():
        shutil.copy2(path, backup)

    if path.suffix.lower() == ".xlsx":
        patch_xlsx(path)
    elif path.suffix.lower() == ".ods":
        patch_ods(path)
    else:
        raise SystemExit(f"Unsupported format: {path.suffix}")

    print(f"Updated {path}")
    print(f"Backup: {backup}")


if __name__ == "__main__":
    main()
